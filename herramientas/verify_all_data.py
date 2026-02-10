from openpyxl import load_workbook
import re

# Read Excel data
wb = load_workbook('piping-dimensions.xlsx')
ws = wb.active

excel_data = {}
for row_idx in range(2, ws.max_row + 1):
    nps = ws.cell(row_idx, 1).value
    od = ws.cell(row_idx, 2).value
    schedule = ws.cell(row_idx, 3).value
    thickness = ws.cell(row_idx, 4).value
    
    if nps is None or od is None or schedule is None or thickness is None:
        continue
    
    schedule_str = str(schedule)
    
    if nps not in excel_data:
        excel_data[nps] = {
            'OD': od,
            'schedules': {}
        }
    
    excel_data[nps]['schedules'][schedule_str] = thickness

# Read HTML data
with open('00_mawp_calculator.html', 'r', encoding='utf-8') as f:
    html_content = f.read()

# Extract pipeDimensions from HTML
# Find the pipeDimensions object
start_marker = 'const pipeDimensions = {'
end_marker = '};'

start_idx = html_content.find(start_marker)
if start_idx == -1:
    print("ERROR: Could not find pipeDimensions in HTML")
    exit(1)

# Find the matching closing brace
brace_count = 0
end_idx = start_idx + len(start_marker)
for i in range(start_idx + len(start_marker), len(html_content)):
    if html_content[i] == '{':
        brace_count += 1
    elif html_content[i] == '}':
        if brace_count == 0:
            end_idx = i + 1
            break
        brace_count -= 1

pipe_dims_str = html_content[start_idx:end_idx]

# Parse the JavaScript object (simplified parsing)
html_data = {}
current_nps = None

for line in pipe_dims_str.split('\n'):
    line = line.strip()
    
    # Match NPS line like "0.5: {"
    nps_match = re.match(r'^([\d.]+):\s*{', line)
    if nps_match:
        current_nps = float(nps_match.group(1))
        html_data[current_nps] = {'OD': None, 'schedules': {}}
        continue
    
    # Match OD line like "OD: 0.84,"
    od_match = re.match(r'OD:\s*([\d.]+)', line)
    if od_match and current_nps is not None:
        html_data[current_nps]['OD'] = float(od_match.group(1))
        continue
    
    # Match schedule line like "STD: 0.109,"
    sched_match = re.match(r'([A-Z0-9]+):\s*([\d.]+)', line)
    if sched_match and current_nps is not None:
        schedule = sched_match.group(1)
        thickness = float(sched_match.group(2))
        html_data[current_nps]['schedules'][schedule] = thickness

print("=" * 100)
print("DETAILED DATA VERIFICATION: Excel vs HTML")
print("=" * 100)

errors_found = 0
warnings_found = 0

for nps in sorted(excel_data.keys()):
    excel_od = excel_data[nps]['OD']
    excel_schedules = excel_data[nps]['schedules']
    
    print(f"\n{'='*100}")
    print(f"NPS {nps}:")
    print(f"{'='*100}")
    
    if nps not in html_data:
        print(f"  ERROR: NPS {nps} NOT FOUND IN HTML!")
        errors_found += 1
        continue
    
    html_od = html_data[nps]['OD']
    html_schedules = html_data[nps]['schedules']
    
    # Check OD
    if abs(excel_od - html_od) > 0.001:
        print(f"  ERROR: OD mismatch!")
        print(f"    Excel: {excel_od}")
        print(f"    HTML:  {html_od}")
        errors_found += 1
    else:
        print(f"  OK: OD = {excel_od}")
    
    # Check schedules
    excel_sched_set = set(excel_schedules.keys())
    html_sched_set = set(html_schedules.keys())
    
    missing_in_html = excel_sched_set - html_sched_set
    extra_in_html = html_sched_set - excel_sched_set
    
    if missing_in_html:
        print(f"  ERROR: Schedules missing in HTML: {sorted(missing_in_html)}")
        errors_found += len(missing_in_html)
    
    if extra_in_html:
        print(f"  WARNING: Extra schedules in HTML (not in Excel): {sorted(extra_in_html)}")
        warnings_found += len(extra_in_html)
    
    # Check thickness values for common schedules
    common_schedules = excel_sched_set & html_sched_set
    thickness_errors = []
    
    for schedule in sorted(common_schedules):
        excel_t = excel_schedules[schedule]
        html_t = html_schedules[schedule]
        
        if abs(excel_t - html_t) > 0.001:
            thickness_errors.append(f"    {schedule}: Excel={excel_t}, HTML={html_t}")
            errors_found += 1
    
    if thickness_errors:
        print(f"  ERROR: Thickness mismatches:")
        for err in thickness_errors:
            print(err)
    else:
        print(f"  OK: All {len(common_schedules)} schedule thicknesses match")

print("\n" + "=" * 100)
print("FINAL SUMMARY")
print("=" * 100)
print(f"Total NPS sizes checked: {len(excel_data)}")
print(f"Errors found: {errors_found}")
print(f"Warnings found: {warnings_found}")

if errors_found == 0 and warnings_found == 0:
    print("\n*** ALL DATA IS CORRECT! ***")
elif errors_found == 0:
    print(f"\n*** NO ERRORS, but {warnings_found} warnings (extra schedules in HTML) ***")
else:
    print(f"\n*** {errors_found} ERRORS NEED TO BE FIXED ***")
