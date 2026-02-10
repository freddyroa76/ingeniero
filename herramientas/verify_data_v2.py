from openpyxl import load_workbook
import json
import re

# Read Excel data
wb = load_workbook('piping-dimensions.xlsx')
ws = wb.active

excel_data = {}
for row_idx in range(2, ws.max_row + 1):
    nps = ws.cell(row_idx, 1).value
    od = ws.cell(row_idx, 2).value
    schedule = ws.cell(row_idx, 3).value
    thickness = ws.cell(row_idx, 4).value
    
    if nps is None or od is None or schedule is None or thickness is None:
        continue
    
    schedule_str = str(schedule)
    
    if nps not in excel_data:
        excel_data[nps] = {
            'OD': od,
            'schedules': {}
        }
    
    excel_data[nps]['schedules'][schedule_str] = thickness

# Read HTML and extract JavaScript object
with open('00_mawp_calculator.html', 'r', encoding='utf-8') as f:
    html_content = f.read()

# Find pipeDimensions
start_marker = 'const pipeDimensions = {'
start_idx = html_content.find(start_marker)
if start_idx == -1:
    print("ERROR: Could not find pipeDimensions")
    exit(1)

# Find matching closing brace
brace_count = 1
i = start_idx + len(start_marker)
while i < len(html_content) and brace_count > 0:
    if html_content[i] == '{':
        brace_count += 1
    elif html_content[i] == '}':
        brace_count -= 1
    i += 1

js_obj = html_content[start_idx + len('const pipeDimensions = '):i]

# Convert JavaScript to JSON
# Replace unquoted property keys with quoted keys, but not numbers
js_obj = re.sub(r'([a-zA-Z_]\w*):', r'"\1":', js_obj)
# Remove trailing commas before closing braces
js_obj = re.sub(r',(\s*[}\]])', r'\1', js_obj)

try:
    html_data_raw = json.loads(js_obj)
    # Convert string keys to float for NPS
    html_data = {}
    for nps_str, data in html_data_raw.items():
        nps_float = float(nps_str)
        html_data[nps_float] = {
            'OD': data['OD'],
            'schedules': data['schedules']
        }
except json.JSONDecodeError as e:
    print(f"ERROR parsing JavaScript: {e}")
    print("First 500 chars of js_obj:")
    print(js_obj[:500])
    exit(1)

print("=" * 100)
print("DETAILED DATA VERIFICATION")
print("=" * 100)

errors = []
warnings = []

for nps in sorted(excel_data.keys()):
    excel_od = excel_data[nps]['OD']
    excel_schedules = excel_data[nps]['schedules']
    
    if nps not in html_data:
        errors.append(f"NPS {nps}: NOT FOUND IN HTML")
        continue
    
    html_od = html_data[nps]['OD']
    html_schedules = html_data[nps]['schedules']
    
    # Check OD
    if abs(excel_od - html_od) > 0.001:
        errors.append(f"NPS {nps}: OD mismatch - Excel={excel_od}, HTML={html_od}")
    
    # Check schedules
    excel_sched_set = set(excel_schedules.keys())
    html_sched_set = set(html_schedules.keys())
    
    missing = excel_sched_set - html_sched_set
    extra = html_sched_set - excel_sched_set
    
    if missing:
        errors.append(f"NPS {nps}: Missing schedules in HTML: {sorted(missing)}")
    
    if extra:
        warnings.append(f"NPS {nps}: Extra schedules in HTML: {sorted(extra)}")
    
    # Check thicknesses
    for schedule in sorted(excel_sched_set & html_sched_set):
        excel_t = excel_schedules[schedule]
        html_t = html_schedules[schedule]
        
        if abs(excel_t - html_t) > 0.001:
            errors.append(f"NPS {nps}, Schedule {schedule}: Thickness mismatch - Excel={excel_t}, HTML={html_t}")

print("\nERRORS:")
if errors:
    for i, err in enumerate(errors, 1):
        print(f"  {i}. {err}")
else:
    print("  None!")

print("\nWARNINGS:")
if warnings:
    for i, warn in enumerate(warnings, 1):
        print(f"  {i}. {warn}")
else:
    print("  None!")

print("\n" + "=" * 100)
print(f"Total errors: {len(errors)}")
print(f"Total warnings: {len(warnings)}")

if len(errors) == 0:
    print("\n*** ALL DATA IS CORRECT! ***")
else:
    print(f"\n*** {len(errors)} ERRORS NEED TO BE FIXED ***")
