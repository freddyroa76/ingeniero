import csv
import sys

# Try to read Excel using xlrd or openpyxl
try:
    from openpyxl import load_workbook
    wb = load_workbook('piping-dimensions.xlsx')
    ws = wb.active
    
    print(f"Total rows: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")
    print("\nHeaders:")
    headers = [cell.value for cell in ws[1]]
    print(" | ".join(str(h) for h in headers))
    
    print("\nFirst 50 data rows:")
    for row_idx in range(2, min(52, ws.max_row + 1)):
        row_data = [cell.value for cell in ws[row_idx]]
        print(" | ".join(str(v) if v is not None else "" for v in row_data))
    
    print(f"\n\nLast 10 rows:")
    for row_idx in range(max(2, ws.max_row - 9), ws.max_row + 1):
        row_data = [cell.value for cell in ws[row_idx]]
        print(" | ".join(str(v) if v is not None else "" for v in row_data))
        
except ImportError:
    print("openpyxl not available, trying xlrd...")
    try:
        import xlrd
        wb = xlrd.open_workbook('piping-dimensions.xlsx')
        ws = wb.sheet_by_index(0)
        
        print(f"Total rows: {ws.nrows}")
        print(f"Total columns: {ws.ncols}")
        print("\nHeaders:")
        headers = ws.row_values(0)
        print(" | ".join(str(h) for h in headers))
        
        print("\nFirst 50 data rows:")
        for row_idx in range(1, min(51, ws.nrows)):
            row_data = ws.row_values(row_idx)
            print(" | ".join(str(v) for v in row_data))
            
        print(f"\n\nLast 10 rows:")
        for row_idx in range(max(1, ws.nrows - 10), ws.nrows):
            row_data = ws.row_values(row_idx)
            print(" | ".join(str(v) for v in row_data))
    except ImportError:
        print("Neither openpyxl nor xlrd available. Please install one of them.")
        sys.exit(1)
