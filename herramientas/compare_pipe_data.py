from openpyxl import load_workbook

wb = load_workbook('piping-dimensions.xlsx')
ws = wb.active

# Extract data from Excel
excel_data = {}
for row_idx in range(2, ws.max_row + 1):
    nps = ws.cell(row_idx, 1).value
    od = ws.cell(row_idx, 2).value
    schedule = ws.cell(row_idx, 3).value
    thickness = ws.cell(row_idx, 4).value
    
    if nps is None or od is None or schedule is None or thickness is None:
        continue
    
    schedule_str = str(schedule)
    
    if nps not in excel_data:
        excel_data[nps] = {
            'OD': od,
            'schedules': {}
        }
    
    excel_data[nps]['schedules'][schedule_str] = thickness

# NPS sizes in HTML (updated after adding missing sizes)
html_nps_sizes = [0.125, 0.25, 0.375, 0.5, 0.75, 1, 1.25, 1.5, 2, 2.5, 3, 3.5, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 30, 32, 34, 36, 42]

print("=" * 80)
print("COMPARISON: Excel vs HTML")
print("=" * 80)

print("\n1. NPS SIZES COMPARISON:")
print(f"   Excel has {len(excel_data)} NPS sizes")
print(f"   HTML has {len(html_nps_sizes)} NPS sizes")

excel_nps_set = set(excel_data.keys())
html_nps_set = set(html_nps_sizes)

missing_in_html = excel_nps_set - html_nps_set
extra_in_html = html_nps_set - excel_nps_set

if missing_in_html:
    print(f"\n   MISSING IN HTML: {sorted(missing_in_html)}")
else:
    print("\n   OK: All Excel NPS sizes are in HTML")

if extra_in_html:
    print(f"   EXTRA IN HTML (not in Excel): {sorted(extra_in_html)}")

print("\n2. DETAILED COMPARISON BY NPS:")
print("-" * 80)

for nps in sorted(excel_data.keys()):
    excel_schedules = set(excel_data[nps]['schedules'].keys())
    excel_od = excel_data[nps]['OD']
    
    print(f"\nNPS {nps}:")
    print(f"  Excel: OD={excel_od}, {len(excel_schedules)} schedules: {sorted(excel_schedules)}")
    
    if nps in html_nps_set:
        print(f"  OK: Present in HTML")
    else:
        print(f"  MISSING IN HTML")

print("\n" + "=" * 80)
print("SUMMARY:")
print("=" * 80)
print(f"Total NPS sizes in Excel: {len(excel_data)}")
print(f"Total NPS sizes in HTML: {len(html_nps_sizes)}")
print(f"Missing in HTML: {len(missing_in_html)}")
print(f"Coverage: {len(html_nps_set & excel_nps_set)}/{len(excel_data)} ({100*len(html_nps_set & excel_nps_set)/len(excel_data):.1f}%)")
