from openpyxl import load_workbook
import json

wb = load_workbook('piping-dimensions.xlsx')
ws = wb.active

# Create a dictionary structure similar to the JavaScript
pipe_data = {}

for row_idx in range(2, ws.max_row + 1):
    nps = ws.cell(row_idx, 1).value
    od = ws.cell(row_idx, 2).value
    schedule = ws.cell(row_idx, 3).value
    thickness = ws.cell(row_idx, 4).value
    
    if nps is None or od is None or schedule is None or thickness is None:
        continue
    
    # Convert schedule to string
    schedule_str = str(schedule)
    
    if nps not in pipe_data:
        pipe_data[nps] = {
            'OD': od,
            'schedules': {}
        }
    
    pipe_data[nps]['schedules'][schedule_str] = thickness

# Print in a format easy to compare
print("Total NPS sizes:", len(pipe_data))
print("\nNPS sizes found:", sorted(pipe_data.keys()))

print("\n\nDetailed data by NPS:")
for nps in sorted(pipe_data.keys()):
    schedules = pipe_data[nps]['schedules']
    print(f"\n{nps}: OD={pipe_data[nps]['OD']}, Schedules: {sorted(schedules.keys())}")
    print(f"  Total schedules: {len(schedules)}")

# Save to JSON for easy comparison
with open('pipe_dimensions_from_excel.json', 'w') as f:
    json.dump(pipe_data, f, indent=2, sort_keys=True)
    
print("\n\nData saved to pipe_dimensions_from_excel.json")
