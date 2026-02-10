from openpyxl import load_workbook

# Read Excel data
wb = load_workbook('piping-dimensions.xlsx')
ws = wb.active

excel_data = {}
for row_idx in range(2, ws.max_row + 1):
    nps = ws.cell(row_idx, 1).value
    od = ws.cell(row_idx, 2).value
    schedule = ws.cell(row_idx, 3).value
    thickness = ws.cell(row_idx, 4).value
    
    if nps is None or od is None or schedule is None or thickness is None:
        continue
    
    schedule_str = str(schedule)
    
    if nps not in excel_data:
        excel_data[nps] = {
            'OD': od,
            'schedules': {}
        }
    
    excel_data[nps]['schedules'][schedule_str] = thickness

# Read HTML and manually extract data
with open('00_mawp_calculator.html', 'r', encoding='utf-8') as f:
    lines = f.readlines()

html_data = {}
current_nps = None
in_schedules = False

for line in lines:
    line = line.strip()
    
    # Match NPS line
    if ': {' in line and 'OD:' not in line and 'schedules:' not in line:
        parts = line.split(':')
        if len(parts) >= 2:
            try:
                nps_val = float(parts[0].strip())
                current_nps = nps_val
                html_data[current_nps] = {'OD': None, 'schedules': {}}
                in_schedules = False
            except:
                pass
    
    # Match OD
    if 'OD:' in line and current_nps is not None:
        parts = line.split('OD:')[1].split(',')[0].strip()
        try:
            html_data[current_nps]['OD'] = float(parts)
        except:
            pass
    
    # Match schedules block
    if 'schedules:' in line:
        in_schedules = True
        # Check if it's inline
        if '{' in line and '}' in line:
            # Inline schedules like: schedules: { STD: 0.226, 40: 0.226, XS: 0.318, 80: 0.318 },
            sched_part = line.split('schedules:')[1].strip()
            sched_part = sched_part.replace('{', '').replace('}', '').replace(',', '')
            pairs = sched_part.split()
            i = 0
            while i < len(pairs) - 1:
                if ':' in pairs[i]:
                    sched_name = pairs[i].replace(':', '')
                    try:
                        thickness = float(pairs[i+1].replace(',', ''))
                        html_data[current_nps]['schedules'][sched_name] = thickness
                        i += 2
                    except:
                        i += 1
                else:
                    i += 1
            in_schedules = False
    
    # Match individual schedule lines
    if in_schedules and ':' in line and current_nps is not None:
        if 'schedules:' not in line and 'OD:' not in line:
            parts = line.split(':')
            if len(parts) >= 2:
                sched_name = parts[0].strip()
                thickness_str = parts[1].split(',')[0].strip()
                try:
                    thickness = float(thickness_str)
                    html_data[current_nps]['schedules'][sched_name] = thickness
                except:
                    pass
    
    # End of schedules block
    if in_schedules and '},' in line and 'schedules' not in line:
        in_schedules = False

print("=" * 100)
print("DETAILED DATA VERIFICATION")
print("=" * 100)

errors = []
warnings = []

for nps in sorted(excel_data.keys()):
    excel_od = excel_data[nps]['OD']
    excel_schedules = excel_data[nps]['schedules']
    
    if nps not in html_data:
        errors.append(f"NPS {nps}: NOT FOUND IN HTML")
        continue
    
    html_od = html_data[nps]['OD']
    html_schedules = html_data[nps]['schedules']
    
    # Check OD
    if html_od is None:
        errors.append(f"NPS {nps}: OD not found in HTML")
    elif abs(excel_od - html_od) > 0.001:
        errors.append(f"NPS {nps}: OD mismatch - Excel={excel_od}, HTML={html_od}")
    
    # Check schedules
    excel_sched_set = set(excel_schedules.keys())
    html_sched_set = set(html_schedules.keys())
    
    missing = excel_sched_set - html_sched_set
    extra = html_sched_set - excel_sched_set
    
    if missing:
        errors.append(f"NPS {nps}: Missing schedules in HTML: {sorted(missing)}")
    
    if extra:
        warnings.append(f"NPS {nps}: Extra schedules in HTML: {sorted(extra)}")
    
    # Check thicknesses
    for schedule in sorted(excel_sched_set & html_sched_set):
        excel_t = excel_schedules[schedule]
        html_t = html_schedules[schedule]
        
        if abs(excel_t - html_t) > 0.001:
            errors.append(f"NPS {nps}, Schedule {schedule}: Thickness mismatch - Excel={excel_t}, HTML={html_t}")

print("\nERRORS:")
if errors:
    for i, err in enumerate(errors, 1):
        print(f"  {i}. {err}")
else:
    print("  None!")

print("\nWARNINGS:")
if warnings:
    for i, warn in enumerate(warnings, 1):
        print(f"  {i}. {warn}")
else:
    print("  None!")

print("\n" + "=" * 100)
print(f"Total errors: {len(errors)}")
print(f"Total warnings: {len(warnings)}")

if len(errors) == 0:
    print("\n*** ALL DATA IS CORRECT! ***")
else:
    print(f"\n*** {len(errors)} ERRORS NEED TO BE FIXED ***")
